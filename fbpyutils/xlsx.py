'''
Functions to read and write MS Excel Spreadsheet files in xls or xlsx formats.
'''
import os
import io
import openpyxl
import xlrd
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
from typing import Union, Dict
import warnings
from fbpyutils.logging import Logger

XLS, XLSX = 0, 1


class ExcelWorkbook:
    """
    Represents an Excel workbook, supporting both .xls and .xlsx formats.

    This class provides a unified interface to read data from Excel files,
    automatically detecting the format and using the appropriate engine
    (openpyxl for .xlsx, xlrd for .xls).

    Attributes:
        workbook: The underlying workbook object from openpyxl or xlrd.
        sheet_names (list[str]): A list of sheet names in the workbook.
        kind (int): The detected file format (XLS or XLSX).
    """

    def __init__(self, xl_file: Union[str, bytes]):
        """
        Initializes an ExcelWorkbook object from a file path or bytes.

        Args:
            xl_file: Path to the Excel file (str) or the file content as bytes.

        Raises:
            FileNotFoundError: If the provided file path does not exist.
            TypeError: If xl_file is not a string path or bytes.
            ValueError: If the file cannot be opened as a valid Excel workbook.

        Example:
            >>> # From a file path
            >>> workbook = ExcelWorkbook('path/to/your/file.xlsx')
            
            >>> # From bytes
            >>> with open('path/to/your/file.xls', 'rb') as f:
            ...     file_bytes = f.read()
            >>> workbook = ExcelWorkbook(file_bytes)
        """
        Logger.debug(f"Initializing ExcelWorkbook with file: {xl_file}")
        data = None

        if isinstance(xl_file, str):
            if os.path.exists(xl_file):
                try:
                    with open(xl_file, 'rb') as f:
                        data = f.read()
                    Logger.info(f"Successfully read Excel file from path: {xl_file}")
                except (OSError, IOError) as e:
                    Logger.error(f"Error reading the Excel file {xl_file}: {e}")
                    raise
            else:
                Logger.error(f"Excel file not found: {xl_file}")
                raise FileNotFoundError(f'File {xl_file} does not exist.')
        elif isinstance(xl_file, bytes):
            data = xl_file
            Logger.info("Received Excel file as bytes.")
        else:
            Logger.error(f"Invalid file reference type: {type(xl_file)}. Must be a file path (str) or bytes.")
            raise TypeError('Invalid file reference. Must be a file path or array of bytes.')

        self.workbook = None
        self.sheet_names = None
        self.kind = XLSX

        try:
            xl_data = io.BytesIO(data)
            warnings.simplefilter("ignore")
            self.workbook = openpyxl.open(xl_data, data_only=True)
            self.sheet_names = self.workbook.sheetnames
            Logger.debug("Workbook opened successfully with openpyxl (XLSX format).")
        except Exception as e_xlsx:
            Logger.warning(f"Failed to open with openpyxl, trying xlrd. Error: {e_xlsx}")
            try:
                xl_data.seek(0)
                self.workbook = xlrd.open_workbook(file_contents=xl_data.read())
                self.sheet_names = self.workbook.sheet_names()
                self.kind = XLS
                Logger.debug("Workbook opened successfully with xlrd (XLS format).")
            except Exception as e_xls:
                Logger.error(f"Failed to open workbook with both openpyxl and xlrd. XLSX error: {e_xlsx}, XLS error: {e_xls}")
                raise ValueError("Could not open workbook. Invalid file format or corrupted file.") from e_xls
        finally:
            warnings.simplefilter("default")
        Logger.info("ExcelWorkbook initialized.")

    def read_sheet(self, sheet_name: str = None) -> tuple[tuple[Union[str, float, int, bool, datetime, None], ...], ...]:
        """
        Reads the contents of a sheet by its name.

        If no sheet name is provided, it reads the first sheet in the workbook.

        Args:
            sheet_name: The name of the sheet to read. Defaults to the first sheet.

        Returns:
            A tuple of tuples, where each inner tuple represents a row of cells.

        Raises:
            NameError: If the specified sheet name does not exist.

        Example:
            >>> workbook = ExcelWorkbook('tests/test_xlsx_file.xlsx')
            >>> # Read the first sheet
            >>> data = workbook.read_sheet()
            >>> # Read a specific sheet by name
            >>> data_sheet2 = workbook.read_sheet('Sheet2')
        """
        sheet_name = sheet_name or self.sheet_names[0]
        Logger.debug(f"Reading sheet: {sheet_name}")

        if sheet_name not in self.sheet_names:
            Logger.error(f"Invalid or nonexistent sheet name: {sheet_name}. Available sheets: {self.sheet_names}")
            raise NameError('Invalid/Nonexistent sheet.')

        try:
            if self.kind == XLSX:
                sh = self.workbook[sheet_name]
                rows = tuple(tuple(c.value for c in r) for r in sh.iter_rows())
            else:  # XLS
                sh = self.workbook.sheet_by_name(sheet_name)
                rows = tuple(tuple(sh.cell_value(r, c) for c in range(sh.ncols)) for r in range(sh.nrows))
            Logger.info(f"Successfully read sheet '{sheet_name}'. Rows read: {len(rows)}")
            return rows
        except Exception as e:
            Logger.error(f"Error reading sheet '{sheet_name}': {e}")
            raise

    def read_sheet_by_index(self, index: int = 0) -> tuple[tuple[Union[str, float, int, bool, datetime, None], ...], ...]:
        """
        Reads the contents of a sheet by its zero-based index.

        Args:
            index: The index of the sheet to read. Defaults to 0 (the first sheet).

        Returns:
            A tuple of tuples, where each inner tuple represents a row of cells.

        Raises:
            IndexError: If the specified sheet index is out of range.

        Example:
            >>> workbook = ExcelWorkbook('tests/test_xlsx_file.xlsx')
            >>> # Read the first sheet (index 0)
            >>> data = workbook.read_sheet_by_index(0)
            >>> # Read the second sheet (index 1)
            >>> data_sheet2 = workbook.read_sheet_by_index(1)
        """
        Logger.debug(f"Reading sheet by index: {index}")

        if not 0 <= index < len(self.sheet_names):
            Logger.error(f"Invalid or nonexistent sheet index: {index}. Total sheets: {len(self.sheet_names)}")
            raise IndexError('Sheet index out of range.')

        return self.read_sheet(self.sheet_names[index])


def get_sheet_names(xl_file: Union[str, bytes]) -> list[str]:
    """
    Retrieves the names of all sheets from an Excel file.

    This is a convenience function that wraps `ExcelWorkbook`.

    Args:
        xl_file: Path to the Excel file or the file content as bytes.

    Returns:
        A list of sheet names.

    Example:
        >>> names = get_sheet_names('tests/test_xlsx_file.xlsx')
        >>> print(names)
        ['Sheet1', 'Sheet2']
    """
    Logger.debug(f"Getting sheet names for file: {xl_file}")
    try:
        xl = ExcelWorkbook(xl_file)
        Logger.info(f"Retrieved sheet names: {xl.sheet_names}")
        return xl.sheet_names
    except Exception as e:
        Logger.error(f"Error getting sheet names from {xl_file}: {e}")
        raise


def get_sheet_by_name(
    xl_file: Union[str, bytes], sheet_name: str
) -> tuple[tuple[Union[str, float, int, bool, datetime, None], ...], ...]:
    """
    Reads a specific sheet from an Excel file by its name.

    This is a convenience function that wraps `ExcelWorkbook`.

    Args:
        xl_file: Path to the Excel file or the file content as bytes.
        sheet_name: The name of the sheet to read.

    Returns:
        The sheet content as a tuple of tuples.

    Example:
        >>> data = get_sheet_by_name('tests/test_xlsx_file.xlsx', 'Sheet1')
    """
    Logger.debug(f"Getting sheet by name '{sheet_name}' from file: {xl_file}")
    try:
        xl = ExcelWorkbook(xl_file)
        sheet_content = xl.read_sheet(sheet_name)
        Logger.info(f"Successfully retrieved sheet '{sheet_name}'.")
        return sheet_content
    except Exception as e:
        Logger.error(f"Error getting sheet by name '{sheet_name}' from {xl_file}: {e}")
        raise


def get_all_sheets(
    xl_file: Union[str, bytes]
) -> Dict[str, tuple[tuple[Union[str, float, int, bool, datetime, None], ...], ...]]:
    """
    Reads all sheets from an Excel file into a dictionary.

    This is a convenience function that wraps `ExcelWorkbook`.

    Args:
        xl_file: Path to the Excel file or the file content as bytes.

    Returns:
        A dictionary where keys are sheet names and values are sheet contents.

    Example:
        >>> all_data = get_all_sheets('tests/test_xlsx_file.xlsx')
        >>> print(all_data.keys())
        dict_keys(['Sheet1', 'Sheet2'])
    """
    Logger.debug(f"Getting all sheets from file: {xl_file}")
    try:
        xl = ExcelWorkbook(xl_file)
        sheet_names = xl.sheet_names
        all_sheets_content = {
            sheet_name: tuple(xl.read_sheet(sheet_name)) for sheet_name in sheet_names
        }
        Logger.info(f"Successfully retrieved all sheets from {xl_file}.")
        return all_sheets_content
    except Exception as e:
        Logger.error(f"Error getting all sheets from {xl_file}: {e}")
        raise


def write_to_sheet(
    df: pd.DataFrame, workbook_path: str, sheet_name: str
) -> None:
    """
    Writes a pandas DataFrame to a specified sheet in an Excel file.

    If the workbook does not exist, it will be created. If the sheet already
    exists in the workbook, a new sheet with an indexed name (e.g., 'Sheet11')
    will be created to avoid overwriting data.

    Args:
        df: The pandas DataFrame to write.
        workbook_path: The path to the target Excel file (.xlsx).
        sheet_name: The desired name for the sheet.

    Example:
        >>> import pandas as pd
        >>> data = {'col1': [1, 2], 'col2': [3, 4]}
        >>> df = pd.DataFrame(data)
        >>>
        >>> # Write to a new file
        >>> write_to_sheet(df, 'output.xlsx', 'MyData')
        >>>
        >>> # Add another sheet to the same file
        >>> df2 = pd.DataFrame({'colA': [5, 6]})
        >>> write_to_sheet(df2, 'output.xlsx', 'MoreData')
        >>>
        >>> # Add a sheet with a name that already exists
        >>> write_to_sheet(df, 'output.xlsx', 'MyData') # Creates 'MyData1'
    """
    Logger.debug(f"Writing DataFrame to Excel sheet '{sheet_name}' in workbook: {workbook_path}")
    try:
        if not os.path.exists(workbook_path):
            Logger.info(f"Workbook does not exist, creating new file: {workbook_path}")
            df.to_excel(
                workbook_path,
                sheet_name=sheet_name,
                index=False,
                freeze_panes=(1, 0),
                header=True,
            )
            Logger.info(f"Successfully created new workbook and wrote sheet '{sheet_name}'.")
        else:
            Logger.info(f"Workbook exists, appending to: {workbook_path}")
            warnings.simplefilter("ignore")
            try:
                # Use 'a' mode to append and 'if_sheet_exists' to handle conflicts.
                # The 'new' option in modern pandas creates a new sheet with an indexed name.
                # The manual check below is kept for robustness across pandas versions.
                with pd.ExcelWriter(workbook_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
                    
                    # Ensure unique sheet name manually for broader compatibility
                    temp_book = load_workbook(workbook_path)
                    current_sheet_names = temp_book.sheetnames
                    temp_book.close()

                    final_sheet_name = sheet_name
                    index = 0
                    while final_sheet_name in current_sheet_names:
                        index += 1
                        final_sheet_name = f"{sheet_name}{index}"
                    
                    Logger.debug(f"Final sheet name determined: {final_sheet_name}")
                    df.to_excel(
                        writer,
                        sheet_name=final_sheet_name,
                        index=False,
                        freeze_panes=(1, 0),
                        header=True,
                    )
                Logger.info(f"Successfully wrote DataFrame to sheet '{final_sheet_name}' in existing workbook.")
            except Exception as e:
                Logger.error(f"Error writing to existing Excel workbook {workbook_path}: {e}")
                raise
            finally:
                warnings.simplefilter("default")
    except Exception as e:
        Logger.error(f"Critical error in write_to_sheet: {e}")
        raise

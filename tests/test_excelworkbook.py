import os
import io
import pytest
import pandas as pd

from openpyxl import load_workbook
from fbpyutils import file as FU
from fbpyutils.xlsx import ExcelWorkbook, get_sheet_names, get_sheet_by_name, get_all_sheets, write_to_sheet


def setup_module(module):
    global root_dir, resources_dir
    root_dir = FU.absolute_path(__file__)
    resources_dir = os.path.sep.join([root_dir, 'resources'])
    print(f"*** SETUP DONE ***")

def teardown_module(module):
    print(f"*** TEARDOWN DONE ***")


@pytest.fixture(scope="module")
def sample_xlsx_file():
    file_path = os.path.sep.join([resources_dir, "sample.xlsx"])
    wb = load_workbook()
    ws = wb.active
    ws.append(["Name", "Age"])
    ws.append(["John Doe", 30])
    ws.append(["Jane Smith", 25])
    wb.save(str(file_path))
    return str(file_path)


def test_excel_workbook_init_with_filepath(sample_xlsx_file):
    file_path = sample_xlsx_file
    workbook = ExcelWorkbook(file_path)
    assert workbook.kind == "XLSX"
    assert workbook.sheet_names == ["Sheet1"]


def test_excel_workbook_init_with_bytes(sample_xlsx_file):
    with open(sample_xlsx_file, "rb") as f:
        data = f.read()
    workbook = ExcelWorkbook(data)
    assert workbook.kind == "XLSX"
    assert workbook.sheet_names == ["Sheet1"]


def test_excel_workbook_init_with_invalid_file():
    with pytest.raises(NameError):
        ExcelWorkbook("path/to/nonexistent.xlsx")


def test_excel_workbook_init_with_invalid_file_type():
    with pytest.raises(TypeError):
        ExcelWorkbook(12345)


def test_excel_workbook_read_sheet(sample_xlsx_file):
    workbook = ExcelWorkbook(sample_xlsx_file)
    rows = workbook.read_sheet("Sheet1")
    assert list(rows) == [("Name", "Age"), ("John Doe", 30), ("Jane Smith", 25)]


def test_excel_workbook_read_sheet_by_index(sample_xlsx_file):
    workbook = ExcelWorkbook(sample_xlsx_file)
    rows = workbook.read_sheet_by_index(0)
    assert list(rows) == [("Name", "Age"), ("John Doe", 30), ("Jane Smith", 25)]


def test_get_sheet_names(sample_xlsx_file):
    sheet_names = get_sheet_names(sample_xlsx_file)
    assert sheet_names == ["Sheet1"]


def test_get_sheet_by_name(sample_xlsx_file):
    rows = get_sheet_by_name(sample_xlsx_file, "Sheet1")
    assert list(rows) == [("Name", "Age"), ("John Doe", 30), ("Jane Smith", 25)]


def test_get_all_sheets(sample_xlsx_file):
    sheets = get_all_sheets(sample_xlsx_file)
    assert sheets == {"Sheet1": [("Name", "Age"), ("John Doe", 30), ("Jane Smith", 25)]}


def test_write_to_sheet(tmp_path):
    df = pd.DataFrame({"Name": ["John Doe", "Jane Smith"], "Age": [30, 25]})
    file_path = tmp_path.joinpath("output.xlsx")
    sheet_name = "Sheet1"
    write_to_sheet(df, str(file_path), sheet_name)
    assert os.path.exists(file_path)
    workbook = load_workbook(file_path)
    assert sheet_name in workbook.sheetnames
    sheet = workbook[sheet_name]
    assert sheet["A1"].value == "Name"
    assert sheet["B1"].value == "Age"
    assert sheet["A2"].value == "John Doe"
    assert sheet["B2"].value == 30
    assert sheet["A3"].value == "Jane Smith"
    assert sheet["B3"].value == 25
    workbook.close()

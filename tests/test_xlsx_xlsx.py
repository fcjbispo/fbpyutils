import pytest
from fbpyutils.xlsx import ExcelWorkbook, XLSX, XLS
import os
import tempfile
from unittest.mock import patch

def test_excel_workbook_constructor_xlsx():
    # Test constructor with XLSX file path
    workbook = ExcelWorkbook('tests/test_xlsx_file.xlsx')
    assert isinstance(workbook, ExcelWorkbook)
    assert workbook.kind == XLSX

def test_excel_workbook_constructor_bytes_xlsx():
    # Test constructor with XLSX file bytes
    with open('tests/test_xlsx_file.xlsx', 'rb') as f:
        file_bytes = f.read()
    workbook = ExcelWorkbook(file_bytes)
    assert isinstance(workbook, ExcelWorkbook)
    assert workbook.kind == XLSX

def test_excel_workbook_read_sheet_xlsx():
    # Test read_sheet method with XLSX file
    workbook = ExcelWorkbook('tests/test_xlsx_file.xlsx')
    sheet_content = workbook.read_sheet()
    assert isinstance(sheet_content, tuple)
    assert len(sheet_content) > 0

def test_excel_workbook_read_sheet_by_index_xlsx():
    # Test read_sheet_by_index method with XLSX file
    workbook = ExcelWorkbook('tests/test_xlsx_file.xlsx')
    sheet_content = workbook.read_sheet_by_index(0)
    assert isinstance(sheet_content, tuple)
    assert len(sheet_content) > 0

def test_get_sheet_names_xlsx():
    # Test get_sheet_names function with XLSX file
    sheet_names = ExcelWorkbook('tests/test_xlsx_file.xlsx').sheet_names
    assert isinstance(sheet_names, list)
    assert len(sheet_names) > 0

def test_get_sheet_by_name_xlsx():
    # Test get_sheet_by_name function with XLSX file
    sheet_name = ExcelWorkbook('tests/test_xlsx_file.xlsx').sheet_names[0]
    sheet_content = ExcelWorkbook('tests/test_xlsx_file.xlsx').read_sheet(sheet_name)
    assert isinstance(sheet_content, tuple)
    assert len(sheet_content) > 0


def test_get_all_sheets_xlsx():
    # Test get_all_sheets function with XLSX file
    from fbpyutils.xlsx import get_all_sheets
    all_sheets = get_all_sheets('tests/test_xlsx_file.xlsx')
    assert isinstance(all_sheets, dict)
    assert len(all_sheets) > 0
def test_write_to_sheet_xlsx():
    # Test write_to_sheet function with XLSX file
    import pandas as pd
    df = pd.DataFrame({'col1': [3, 4], 'col2': ['c', 'd']})
    workbook_path = 'tests/test_write_xlsx_file.xlsx'
    sheet_name = 'Sheet1'
    from fbpyutils.xlsx import write_to_sheet
    write_to_sheet(df, workbook_path, sheet_name)
    assert os.path.exists(workbook_path)
    workbook = ExcelWorkbook(workbook_path)
    assert sheet_name in workbook.sheet_names
    os.remove(workbook_path) # Cleanup test file

def test_excel_workbook_constructor_file_not_found():
    # Test constructor with a non-existent file path
    with pytest.raises(FileNotFoundError, match="File non_existent_file.xlsx does not exist."):
        ExcelWorkbook('non_existent_file.xlsx')

def test_excel_workbook_constructor_invalid_type():
    # Test constructor with an invalid file type
    with pytest.raises(TypeError, match="Invalid file reference. Must be a file path or array of bytes."):
        ExcelWorkbook(12345)

@patch('builtins.open', side_effect=IOError("Simulated IOError"))
def test_excel_workbook_constructor_io_error(mock_open):
    # Test constructor with an IOError during file reading
    with pytest.raises(IOError, match="Simulated IOError"):
        ExcelWorkbook('tests/test_xlsx_file.xlsx')

def test_excel_workbook_constructor_xls():
    # Test constructor with XLS file path (mocking xlrd)
    # Create a dummy xls-like content for xlrd to open
    # This is a simplified representation; real XLS is binary
    dummy_xls_content = b"dummy xls content"
    with patch('xlrd.open_workbook') as mock_xlrd_open_workbook, \
         patch('openpyxl.open', side_effect=OSError("Simulated OSError to trigger xlrd path")) as mock_openpyxl_open: # Ensure openpyxl fails first
        
        mock_workbook_xlrd = mock_xlrd_open_workbook.return_value
        mock_workbook_xlrd.sheet_names.return_value = ['Sheet1_xls']

        workbook = ExcelWorkbook(dummy_xls_content)
        assert isinstance(workbook, ExcelWorkbook)
        assert workbook.kind == XLS
        assert workbook.sheet_names == ['Sheet1_xls']
        mock_xlrd_open_workbook.assert_called_once()
        # Check that openpyxl.open was called before xlrd.open_workbook
        assert mock_openpyxl_open.called

def test_excel_workbook_read_sheet_invalid_name_xlsx():
    # Test read_sheet with an invalid sheet name for XLSX
    workbook = ExcelWorkbook('tests/test_xlsx_file.xlsx')
    with pytest.raises(NameError, match="Invalid/Nonexistent sheet."):
        workbook.read_sheet('NonExistentSheet')

def test_excel_workbook_read_sheet_by_index_invalid_index_xlsx():
    # Test read_sheet_by_index with an invalid index for XLSX
    workbook = ExcelWorkbook('tests/test_xlsx_file.xlsx')
    with pytest.raises(NameError, match="Invalid/Nonexistent sheet."):
        workbook.read_sheet_by_index(100) # Assuming there are less than 100 sheets

@patch('fbpyutils.xlsx.ExcelWorkbook', side_effect=ValueError("Simulated ExcelWorkbook init error"))
def test_get_sheet_names_workbook_init_error(mock_excel_workbook):
    from fbpyutils.xlsx import get_sheet_names
    with pytest.raises(ValueError, match="Simulated ExcelWorkbook init error"):
        get_sheet_names('dummy_path.xlsx')

@patch('fbpyutils.xlsx.ExcelWorkbook', side_effect=ValueError("Simulated ExcelWorkbook init error"))
def test_get_sheet_by_name_workbook_init_error(mock_excel_workbook):
    from fbpyutils.xlsx import get_sheet_by_name
    with pytest.raises(ValueError, match="Simulated ExcelWorkbook init error"):
        get_sheet_by_name('dummy_path.xlsx', 'Sheet1')

@patch('fbpyutils.xlsx.ExcelWorkbook', side_effect=ValueError("Simulated ExcelWorkbook init error"))
def test_get_all_sheets_workbook_init_error(mock_excel_workbook):
    from fbpyutils.xlsx import get_all_sheets
    with pytest.raises(ValueError, match="Simulated ExcelWorkbook init error"):
        get_all_sheets('dummy_path.xlsx')

def test_write_to_sheet_existing_workbook_new_sheet_xlsx():
    # Test write_to_sheet to an existing workbook with a new sheet name
    import pandas as pd
    from fbpyutils.xlsx import write_to_sheet

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmpfile:
        workbook_path = tmpfile.name
    
    try:
        # Create an initial valid xlsx workbook using openpyxl
        from openpyxl import Workbook as OpenpyxlWorkbook
        wb_initial = OpenpyxlWorkbook()
        wb_initial.save(workbook_path)

        # Create an initial workbook sheet using our function
        df1 = pd.DataFrame({'colA': [1, 2], 'colB': ['a', 'b']})
        write_to_sheet(df1, workbook_path, "OriginalSheet")

        # Write a new sheet to the existing workbook
        df2 = pd.DataFrame({'col1': [3, 4], 'col2': ['c', 'd']})
        new_sheet_name = "NewSheet"
        write_to_sheet(df2, workbook_path, new_sheet_name)

        wb = ExcelWorkbook(workbook_path)
        assert "OriginalSheet" in wb.sheet_names
        assert new_sheet_name in wb.sheet_names
        
        # Verify content of the new sheet
        new_sheet_content = wb.read_sheet(new_sheet_name)
        assert new_sheet_content[0] == ('col1', 'col2') # Header
        assert new_sheet_content[1] == (3, 'c')
        assert new_sheet_content[2] == (4, 'd')

    finally:
        if os.path.exists(workbook_path):
            os.remove(workbook_path)

def test_write_to_sheet_existing_workbook_duplicate_sheet_name_xlsx():
    # Test write_to_sheet to an existing workbook with a duplicate sheet name
    import pandas as pd
    from fbpyutils.xlsx import write_to_sheet

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmpfile:
        workbook_path = tmpfile.name

    try:
        # Create an initial valid xlsx workbook using openpyxl
        from openpyxl import Workbook as OpenpyxlWorkbook
        wb_initial = OpenpyxlWorkbook()
        wb_initial.save(workbook_path)
        
        # Create an initial workbook sheet using our function
        df1 = pd.DataFrame({'colA': [10, 20], 'colB': ['x', 'y']})
        original_sheet_name = "DuplicateSheet"
        write_to_sheet(df1, workbook_path, original_sheet_name)

        # Write another sheet with the same name
        df2 = pd.DataFrame({'colX': [30, 40], 'colY': ['p', 'q']})
        write_to_sheet(df2, workbook_path, original_sheet_name) # Attempt to write with duplicate name

        wb = ExcelWorkbook(workbook_path)
        # Check that the original sheet exists
        assert original_sheet_name in wb.sheet_names
        # Check that a new sheet with an incremented name was created
        assert f"{original_sheet_name}1" in wb.sheet_names 

        # Write a third sheet with the same original name
        df3 = pd.DataFrame({'colZ': [50, 60]})
        write_to_sheet(df3, workbook_path, original_sheet_name)
        
        wb_updated = ExcelWorkbook(workbook_path)
        assert original_sheet_name in wb_updated.sheet_names
        assert f"{original_sheet_name}1" in wb_updated.sheet_names
        assert f"{original_sheet_name}2" in wb_updated.sheet_names # Should be SheetName2 now

    finally:
        if os.path.exists(workbook_path):
            os.remove(workbook_path)

def test_write_to_sheet_empty_dataframe_xlsx():
    # Test write_to_sheet with an empty DataFrame
    import pandas as pd
    from fbpyutils.xlsx import write_to_sheet

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmpfile:
        workbook_path = tmpfile.name
    sheet_name = "EmptySheet"
    
    try:
        # Create an initial valid xlsx workbook using openpyxl
        from openpyxl import Workbook as OpenpyxlWorkbook
        wb_initial = OpenpyxlWorkbook()
        # Ensure there's at least one sheet for openpyxl to save, even if it's just the default one.
        # Our write_to_sheet should handle adding a new one or replacing.
        # If the goal is to test writing an empty df to a TRULY empty (no sheets) workbook,
        # that scenario might not be supported by openpyxl's save or pandas' to_excel.
        # For this test, let's ensure the temp file is a valid (but minimal) xlsx.
        if not wb_initial.sheetnames: # If, for some reason, it was created without any sheets
            wb_initial.create_sheet("DefaultSheetForTest") # Add a default sheet
        elif "Sheet" in wb_initial.sheetnames and len(wb_initial.sheetnames) > 1: # If default 'Sheet' exists and others too
             wb_initial.remove(wb_initial["Sheet"]) # Remove default only if other sheets exist
        elif "Sheet" in wb_initial.sheetnames and len(wb_initial.sheetnames) == 1:
            pass # Keep the default 'Sheet' if it's the only one.
        wb_initial.save(workbook_path)

        df_empty = pd.DataFrame()
        write_to_sheet(df_empty, workbook_path, sheet_name)
        
        assert os.path.exists(workbook_path)
        wb = ExcelWorkbook(workbook_path)
        assert sheet_name in wb.sheet_names
        sheet_content = wb.read_sheet(sheet_name)
        # Depending on pandas version, it might write an empty sheet or a sheet with just headers if columns are defined
        # For a truly empty DataFrame (no columns), openpyxl might create a sheet with 1 empty row or just the sheet.
        # Let's assert the sheet exists and is essentially empty or has minimal structure.
        assert len(sheet_content) <= 1 # Allow for a header row or completely empty
        if sheet_content: # if there's any row (e.g. header for df with columns but no rows)
             assert all(c is None for c in sheet_content[0]) or not sheet_content[0]


    finally:
        if os.path.exists(workbook_path):
            os.remove(workbook_path)

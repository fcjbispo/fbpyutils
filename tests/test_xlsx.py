import os
import pytest
import logging
import pandas as pd

import openpyxl
from fbpyutils import file as FU
from fbpyutils.xlsx import ExcelWorkbook, XLSX

from openpyxl import load_workbook
from fbpyutils.xlsx import write_to_sheet


def setup_module(module):
    global root_dir, resources_dir
    root_dir = FU.absolute_path(__file__)
    resources_dir = os.path.sep.join([root_dir, "resources"])
    if not os.path.exists(resources_dir):
        os.makedirs(resources_dir, exist_ok=True)
    logging.debug(f"*** SETUP DONE ***")


def teardown_module(module):
    logging.debug(f"*** TEARDOWN DONE ***")


@pytest.fixture
def test_file():
    # create a test file
    test_data = (
        ("Name", "Age", "Gender"),
        ("John", 25, "Male"),
        ("Jane", 30, "Female"),
        ("Bob", 40, "Male"),
    )
    test_file = os.path.sep.join([resources_dir, "sample.xlsx"])
    if os.path.exists(test_file):
        os.remove(test_file)
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in test_data:
        ws.append(row)
    wb.save(test_file)
    yield test_file
    os.remove(test_file)


def test_init_with_path(test_file):
    # test initializing ExcelWorkbook with file path
    wb = ExcelWorkbook(test_file)
    assert wb.sheet_names == ["Sheet"]
    assert wb.kind == XLSX


def test_init_with_bytes(test_file):
    # test initializing ExcelWorkbook with bytes
    with open(test_file, "rb") as f:
        data = f.read()
        f.close()

    wb = ExcelWorkbook(data)
    assert wb.sheet_names == ["Sheet"]
    assert wb.kind == XLSX


def test_read_sheet_by_name(test_file):
    # test reading sheet by name
    wb = ExcelWorkbook(test_file)
    data = wb.read_sheet("Sheet")
    assert tuple(data) == (
        ("Name", "Age", "Gender"),
        ("John", 25, "Male"),
        ("Jane", 30, "Female"),
        ("Bob", 40, "Male"),
    )


def test_read_sheet_by_index(test_file):
    # test reading sheet by index
    wb = ExcelWorkbook(test_file)
    data = wb.read_sheet_by_index(0)
    assert tuple(data) == (
        ("Name", "Age", "Gender"),
        ("John", 25, "Male"),
        ("Jane", 30, "Female"),
        ("Bob", 40, "Male"),
    )


def test_read_sheet_invalid_sheet(test_file):
    # test reading invalid sheet
    wb = ExcelWorkbook(test_file)
    with pytest.raises(NameError):
        wb.read_sheet("Invalid Sheet")


def test_read_sheet_by_index_invalid_sheet(test_file):
    # test reading invalid sheet by index
    wb = ExcelWorkbook(test_file)
    with pytest.raises(NameError):
        wb.read_sheet_by_index(1)


def test_init_invalid_file():
    # test initializing ExcelWorkbook with invalid file
    with pytest.raises(NameError):
        wb = ExcelWorkbook("invalid_file.xlsx")


def test_init_invalid_file_type():
    # test initializing ExcelWorkbook with invalid file type
    with pytest.raises(TypeError):
        wb = ExcelWorkbook(123)


def test_write_to_sheet_with_new_workbook(tmp_path):
    # Create a new workbook and write a DataFrame to it
    df = pd.DataFrame({"A": [1, 2, 3], "B": [4, 5, 6]})
    workbook_path = os.path.join(tmp_path, "test.xlsx")
    sheet_name = "Sheet1"
    write_to_sheet(df, workbook_path, sheet_name)
    # Check that the workbook file exists and contains the expected data
    assert os.path.exists(workbook_path)
    book = load_workbook(workbook_path)
    sheet = book[sheet_name]
    assert sheet["A1"].value == "A"
    assert sheet["B1"].value == "B"
    assert sheet["A2"].value == 1
    assert sheet["B2"].value == 4
    assert sheet["A3"].value == 2
    assert sheet["B3"].value == 5
    assert sheet["A4"].value == 3
    assert sheet["B4"].value == 6
    book.close()
